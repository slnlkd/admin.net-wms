import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
import os
from datetime import datetime

class SQLServerSchemaExporter:
    def __init__(self, server, database, username=None, password=None, trusted_connection=True):
        """
        SQL Server数据库结构导出器
        
        Args:
            server: 服务器地址
            database: 数据库名称
            username: 用户名（Windows验证时可空）
            password: 密码（Windows验证时可空）
            trusted_connection: 是否使用Windows身份验证
        """
        self.server = server
        self.database = database
        self.username = username
        self.password = password
        self.trusted_connection = trusted_connection
        self.connection = None
        self.table_descriptions = {}  # 存储表描述信息
        
    def connect(self):
        """建立数据库连接"""
        try:
            if self.trusted_connection:
                # Windows身份验证
                conn_str = (
                    f'DRIVER={{ODBC Driver 17 for SQL Server}};'
                    f'SERVER={self.server};'
                    f'DATABASE={self.database};'
                    'Trusted_Connection=yes;'
                )
            else:
                # SQL Server身份验证
                conn_str = (
                    f'DRIVER={{ODBC Driver 17 for SQL Server}};'
                    f'SERVER={self.server};'
                    f'DATABASE={self.database};'
                    f'UID={self.username};'
                    f'PWD={self.password};'
                    'TrustServerCertificate=yes;'
                )
            
            self.connection = pyodbc.connect(conn_str)
            print(f"成功连接到数据库: {self.database}")
            return True
            
        except Exception as e:
            print(f"连接失败: {str(e)}")
            return False

    def execute_query(self, query):
        """执行查询并返回DataFrame，避免pandas警告"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(query)
            
            # 获取列名
            columns = [column[0] for column in cursor.description]
            
            # 获取数据
            data = cursor.fetchall()
            
            # 创建DataFrame
            df = pd.DataFrame.from_records(data, columns=columns)
            return df
            
        except Exception as e:
            print(f"查询执行失败: {str(e)}")
            return None

    def get_table_descriptions(self):
        """获取所有表的描述信息"""
        query = """
        SELECT 
            SCHEMA_NAME(t.schema_id) AS SchemaName,
            t.name AS TableName,
            ISNULL(ep.value, '') AS TableDescription
        FROM sys.tables t
        LEFT JOIN sys.extended_properties ep ON t.object_id = ep.major_id 
            AND ep.minor_id = 0 
            AND ep.name = 'MS_Description'
        WHERE t.is_ms_shipped = 0
        """
        
        try:
            df = self.execute_query(query)
            if df is not None:
                for _, row in df.iterrows():
                    key = f"{row.SchemaName}.{row.TableName}"
                    self.table_descriptions[key] = row.TableDescription if row.TableDescription else row.TableName
                return True
            return False
        except Exception as e:
            print(f"获取表描述失败: {str(e)}")
            return False

    def get_table_schema(self):
        """
        获取详细的表结构信息
        """
        if not self.connection:
            print("请先建立数据库连接")
            return None
        
        query = """
        SELECT 
            SCHEMA_NAME(t.schema_id) AS SchemaName,
            t.name AS TableName,
            c.name AS ColumnName,
            ty.name AS DataType,
            c.max_length AS MaxLength,
            c.precision AS Precision,
            c.scale AS Scale,
            CASE WHEN c.is_nullable = 1 THEN '是' ELSE '否' END AS IsNullable,
            CASE WHEN c.is_identity = 1 THEN '是' ELSE '否' END AS IsIdentity,
            CASE WHEN pk.column_id IS NOT NULL THEN '是' ELSE '否' END AS IsPrimaryKey,
            CASE WHEN fk.parent_column_id IS NOT NULL THEN '是' ELSE '否' END AS IsForeignKey,
            OBJECT_DEFINITION(c.default_object_id) AS DefaultValue,
            ISNULL(ep.value, '') AS Description,
            -- 计算完整的数据类型显示
            CASE 
                WHEN ty.name IN ('varchar', 'char', 'nvarchar', 'nchar') THEN 
                    ty.name + '(' + CASE WHEN c.max_length = -1 THEN 'MAX' ELSE CAST(c.max_length AS VARCHAR) END + ')'
                WHEN ty.name IN ('decimal', 'numeric') THEN 
                    ty.name + '(' + CAST(c.precision AS VARCHAR) + ',' + CAST(c.scale AS VARCHAR) + ')'
                WHEN ty.name IN ('float') THEN 
                    ty.name + '(' + CAST(c.precision AS VARCHAR) + ')'
                ELSE ty.name
            END AS FullDataType
        FROM sys.tables t
        INNER JOIN sys.columns c ON t.object_id = c.object_id
        INNER JOIN sys.types ty ON c.user_type_id = ty.user_type_id
        LEFT JOIN sys.extended_properties ep ON c.object_id = ep.major_id AND c.column_id = ep.minor_id AND ep.minor_id > 0
        LEFT JOIN (
            SELECT ic.object_id, ic.column_id
            FROM sys.index_columns ic
            INNER JOIN sys.indexes i ON ic.object_id = i.object_id AND ic.index_id = i.index_id
            WHERE i.is_primary_key = 1
        ) pk ON c.object_id = pk.object_id AND c.column_id = pk.column_id
        LEFT JOIN sys.foreign_key_columns fk ON c.object_id = fk.parent_object_id AND c.column_id = fk.parent_column_id
        WHERE t.is_ms_shipped = 0
        ORDER BY SCHEMA_NAME(t.schema_id), t.name, c.column_id
        """
        
        return self.execute_query(query)

    def get_table_info(self):
        """
        获取表的基本信息
        """
        query = """
        SELECT 
            SCHEMA_NAME(t.schema_id) AS SchemaName,
            t.name AS TableName,
            p.rows AS RowCounts,
            SUM(a.total_pages) * 8 AS TotalSpaceKB, 
            SUM(a.used_pages) * 8 AS UsedSpaceKB, 
            (SUM(a.total_pages) - SUM(a.used_pages)) * 8 AS UnusedSpaceKB,
            ISNULL(ep.value, '') AS TableDescription,
            t.create_date AS CreateDate,
            t.modify_date AS ModifyDate
        FROM sys.tables t
        INNER JOIN sys.indexes i ON t.object_id = i.object_id
        INNER JOIN sys.partitions p ON i.object_id = p.object_id AND i.index_id = p.index_id
        INNER JOIN sys.allocation_units a ON p.partition_id = a.container_id
        LEFT JOIN sys.extended_properties ep ON t.object_id = ep.major_id AND ep.minor_id = 0 AND ep.name = 'MS_Description'
        WHERE t.is_ms_shipped = 0 AND i.object_id > 255 
        GROUP BY SCHEMA_NAME(t.schema_id), t.name, p.rows, ep.value, t.create_date, t.modify_date
        ORDER BY SCHEMA_NAME(t.schema_id), t.name
        """
        
        df = self.execute_query(query)
        # 更新表描述信息
        if df is not None:
            for _, row in df.iterrows():
                key = f"{row.SchemaName}.{row.TableName}"
                self.table_descriptions[key] = row.TableDescription if row.TableDescription else row.TableName
        return df

    def export_to_excel(self, output_file=None):
        """
        导出表结构到Excel文件
        """
        if not self.connection:
            print("请先建立数据库连接")
            return None
            
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"SQLServer_Schema_Export_{self.database}_{timestamp}.xlsx"
        
        print("开始导出数据库表结构...")
        
        # 获取表描述信息
        self.get_table_descriptions()
        
        # 获取数据
        schema_df = self.get_table_schema()
        table_info_df = self.get_table_info()
        
        if schema_df is None or table_info_df is None:
            print("获取数据失败，无法导出")
            return None
        
        # 创建Excel工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        # 定义样式
        self._define_styles(wb)
        
        # 创建工作表
        self._create_table_info_sheet(wb, table_info_df)
        self._create_data_dictionary_sheet(wb, schema_df)
        
        # 为每个表创建详细工作表
        schemas = schema_df['SchemaName'].unique()
        for schema in schemas:
            schema_tables = schema_df[schema_df['SchemaName'] == schema]['TableName'].unique()
            for table in schema_tables:
                table_data = schema_df[
                    (schema_df['SchemaName'] == schema) & 
                    (schema_df['TableName'] == table)
                ]
                self._create_table_detail_sheet(wb, schema, table, table_data)
        
        # 保存文件
        try:
            wb.save(output_file)
            print(f"导出成功! 文件保存为: {output_file}")
            
            # 显示统计信息
            table_count = len(table_info_df)
            column_count = len(schema_df)
            print(f"统计信息: {table_count} 个表, {column_count} 个列")
            
            return output_file
        except Exception as e:
            print(f"保存Excel文件失败: {str(e)}")
            return None

    def _define_styles(self, wb):
        """定义样式"""
        # 标题样式
        wb.title_style = Font(
            name='微软雅黑',
            size=14,
            bold=True,
            color="2F5496"
        )
        
        # 表头样式
        wb.header_fill = PatternFill(
            start_color="2F5496",
            end_color="2F5496",
            fill_type="solid"
        )
        wb.header_font = Font(
            name='微软雅黑',
            size=10,
            bold=True,
            color="FFFFFF"
        )
        
        # 边框样式
        wb.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 特殊行样式
        wb.primary_key_fill = PatternFill(
            start_color="E2EFDA",
            end_color="E2EFDA",
            fill_type="solid"
        )
        wb.foreign_key_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )
        wb.identity_fill = PatternFill(
            start_color="DDEBF7",
            end_color="DDEBF7",
            fill_type="solid"
        )

    def _create_table_info_sheet(self, wb, table_info_df):
        """创建表信息总览工作表"""
        ws = wb.create_sheet("📊 表信息总览")
        
        # 标题
        ws.cell(row=1, column=1, value="数据库表信息总览").font = wb.title_style
        ws.merge_cells('A1:I1')
        ws.cell(row=2, column=1, value=f"数据库: {self.database} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(size=10, italic=True)
        ws.merge_cells('A2:I2')
        
        # 表头
        headers = ['架构名', '表名', '记录数', '总空间(KB)', '已用空间(KB)', '未用空间(KB)', 
                  '表描述', '创建时间', '修改时间']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = wb.header_font
            cell.fill = wb.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = wb.thin_border
        
        # 添加数据
        for row_idx, row in enumerate(table_info_df.itertuples(), 5):
            ws.cell(row=row_idx, column=1, value=row.SchemaName).border = wb.thin_border
            ws.cell(row=row_idx, column=2, value=row.TableName).border = wb.thin_border
            ws.cell(row=row_idx, column=3, value=row.RowCounts).border = wb.thin_border
            ws.cell(row=row_idx, column=4, value=row.TotalSpaceKB).border = wb.thin_border
            ws.cell(row=row_idx, column=5, value=row.UsedSpaceKB).border = wb.thin_border
            ws.cell(row=row_idx, column=6, value=row.UnusedSpaceKB).border = wb.thin_border
            ws.cell(row=row_idx, column=7, value=row.TableDescription).border = wb.thin_border
            ws.cell(row=row_idx, column=8, value=row.CreateDate.strftime("%Y-%m-%d %H:%M:%S") if row.CreateDate else "").border = wb.thin_border
            ws.cell(row=row_idx, column=9, value=row.ModifyDate.strftime("%Y-%m-%d %H:%M:%S") if row.ModifyDate else "").border = wb.thin_border
        
        # 添加数据条
        try:
            ws.conditional_formatting.add(f'E5:E{4+len(table_info_df)}', 
                DataBarRule(start_type='min', end_type='max', color="638EC6"))
        except:
            pass
        
        self._auto_adjust_columns(ws)
        
        # 冻结窗格
        ws.freeze_panes = 'A5'

    def _create_table_detail_sheet(self, wb, schema, table, table_data):
        """创建表的详细结构工作表"""
        # 获取表描述
        table_key = f"{schema}.{table}"
        table_description = self.table_descriptions.get(table_key, table)
        
        # 创建工作表名称（带描述）
        sheet_name = f"{schema}.{table}-{table_description}"[:31]
        ws = wb.create_sheet(sheet_name)
        
        # 标题
        title_cell = ws.cell(row=1, column=1, value=f"表结构详情: {schema}.{table} - {table_description}")
        title_cell.font = wb.title_style
        ws.merge_cells('A1:L1')
        
        # 基本信息
        ws.cell(row=2, column=1, value=f"架构: {schema}").font = Font(bold=True)
        ws.cell(row=2, column=3, value=f"表名: {table}").font = Font(bold=True)
        ws.cell(row=2, column=5, value=f"列数: {len(table_data)}").font = Font(bold=True)
        ws.cell(row=2, column=7, value=f"表描述: {table_description}").font = Font(bold=True)
        
        # 表头
        headers = ['列名', '数据类型', '完整类型', '最大长度', '精度', '小数位数', 
                  '允许空值', '是否自增', '是否主键', '是否外键', '默认值', '描述']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = wb.header_font
            cell.fill = wb.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = wb.thin_border
        
        # 添加数据
        for row_idx, row in enumerate(table_data.itertuples(), 5):
            # 基础数据
            ws.cell(row=row_idx, column=1, value=row.ColumnName).border = wb.thin_border
            ws.cell(row=row_idx, column=2, value=row.DataType).border = wb.thin_border
            ws.cell(row=row_idx, column=3, value=row.FullDataType).border = wb.thin_border
            ws.cell(row=row_idx, column=4, value=row.MaxLength if row.MaxLength != -1 else "MAX").border = wb.thin_border
            ws.cell(row=row_idx, column=5, value=row.Precision if row.Precision else "").border = wb.thin_border
            ws.cell(row=row_idx, column=6, value=row.Scale if row.Scale else "").border = wb.thin_border
            ws.cell(row=row_idx, column=7, value=row.IsNullable).border = wb.thin_border
            ws.cell(row=row_idx, column=8, value=row.IsIdentity).border = wb.thin_border
            ws.cell(row=row_idx, column=9, value=row.IsPrimaryKey).border = wb.thin_border
            ws.cell(row=row_idx, column=10, value=row.IsForeignKey).border = wb.thin_border
            ws.cell(row=row_idx, column=11, value=row.DefaultValue if row.DefaultValue else "").border = wb.thin_border
            ws.cell(row=row_idx, column=12, value=row.Description).border = wb.thin_border
            
            # 特殊行背景色
            if row.IsPrimaryKey == '是':
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = wb.primary_key_fill
            elif row.IsForeignKey == '是':
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = wb.foreign_key_fill
            elif row.IsIdentity == '是':
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).fill = wb.identity_fill
        
        # 自动调整列宽
        self._auto_adjust_columns(ws)
        
        # 设置指定列的固定宽度为9（约1.87cm）
        fixed_width_columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']  # 最大长度到默认值列
        for col_letter in fixed_width_columns:
            ws.column_dimensions[col_letter].width = 9
        
        # 冻结窗格
        ws.freeze_panes = 'A5'

    def _create_data_dictionary_sheet(self, wb, schema_df):
        """创建数据字典总览工作表"""
        ws = wb.create_sheet("📋 数据字典总览")
        
        # 标题
        ws.cell(row=1, column=1, value="数据字典总览").font = wb.title_style
        ws.merge_cells('A1:J1')
        ws.cell(row=2, column=1, value=f"数据库: {self.database} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(size=10, italic=True)
        ws.merge_cells('A2:J2')
        
        # 表头
        headers = ['架构名', '表名', '表描述', '列名', '完整数据类型', '允许空值', 
                  '是否主键', '是否外键', '默认值', '列描述']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = wb.header_font
            cell.fill = wb.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = wb.thin_border
        
        # 添加数据
        current_table = None
        start_row = 5
        
        for row_idx, row in enumerate(schema_df.itertuples(), 5):
            table_key = f"{row.SchemaName}.{row.TableName}"
            table_description = self.table_descriptions.get(table_key, row.TableName)
            
            # 如果表名发生变化，合并上一个表的行
            if current_table != table_key:
                if current_table is not None and start_row < row_idx - 1:
                    # 合并架构名、表名和表描述列
                    ws.merge_cells(f'A{start_row}:A{row_idx-1}')  # 架构名
                    ws.merge_cells(f'B{start_row}:B{row_idx-1}')  # 表名
                    ws.merge_cells(f'C{start_row}:C{row_idx-1}')  # 表描述
                
                current_table = table_key
                start_row = row_idx
            
            # 添加数据
            ws.cell(row=row_idx, column=1, value=row.SchemaName).border = wb.thin_border
            ws.cell(row=row_idx, column=2, value=row.TableName).border = wb.thin_border
            ws.cell(row=row_idx, column=3, value=table_description).border = wb.thin_border
            ws.cell(row=row_idx, column=4, value=row.ColumnName).border = wb.thin_border
            ws.cell(row=row_idx, column=5, value=row.FullDataType).border = wb.thin_border
            ws.cell(row=row_idx, column=6, value=row.IsNullable).border = wb.thin_border
            ws.cell(row=row_idx, column=7, value=row.IsPrimaryKey).border = wb.thin_border
            ws.cell(row=row_idx, column=8, value=row.IsForeignKey).border = wb.thin_border
            ws.cell(row=row_idx, column=9, value=row.DefaultValue if row.DefaultValue else "").border = wb.thin_border
            ws.cell(row=row_idx, column=10, value=row.Description).border = wb.thin_border
            
            # 特殊行背景色
            if row.IsPrimaryKey == '是':
                for col in range(1, 11):
                    ws.cell(row=row_idx, column=col).fill = wb.primary_key_fill
            elif row.IsForeignKey == '是':
                for col in range(1, 11):
                    ws.cell(row=row_idx, column=col).fill = wb.foreign_key_fill
        
        # 合并最后一个表的行
        if current_table is not None and start_row < len(schema_df) + 4:
            ws.merge_cells(f'A{start_row}:A{len(schema_df)+4}')  # 架构名
            ws.merge_cells(f'B{start_row}:B{len(schema_df)+4}')  # 表名
            ws.merge_cells(f'C{start_row}:C{len(schema_df)+4}')  # 表描述
        
        self._auto_adjust_columns(ws)
        
        # 冻结窗格
        ws.freeze_panes = 'A5'

    def _auto_adjust_columns(self, ws):
        """自动调整列宽 - 修复合并单元格问题"""
        for column_cells in ws.columns:
            # 跳过空列
            if not column_cells:
                continue
                
            # 获取列字母
            column_letter = None
            for cell in column_cells:
                if hasattr(cell, 'column_letter') and not cell.coordinate.startswith('A1'):  # 跳过可能合并的标题单元格
                    column_letter = cell.column_letter
                    break
            
            if not column_letter:
                continue
                
            # 检查是否是固定宽度列，如果是则跳过
            fixed_width_columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            if column_letter in fixed_width_columns:
                continue
                
            # 计算最大长度
            max_length = 0
            for cell in column_cells:
                # 跳过合并的单元格
                if cell.coordinate in ws.merged_cells:
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置列宽
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def close(self):
        """关闭数据库连接"""
        if self.connection:
            self.connection.close()
            print("数据库连接已关闭")

def main():
    """主函数 - 使用示例"""
    # 数据库连接配置 - 两种方式选择一种
    
    # 方式1: Windows身份验证（推荐）
    # config = {
    #     'server': 'localhost',  # 替换为你的服务器地址
    #     'database': 'JC44_WMS',  # 替换为你的数据库名称
    #     'trusted_connection': True,  # 使用Windows身份验证
    #     # 不需要提供用户名和密码
    # }
    
    # 方式2: SQL Server身份验证
    config = {
        'server': '47.95.120.53',
        'database': 'JC44_WMS',
        'username': 'sa',
        'password': 'boxline!@#',
        'trusted_connection': False,  # 使用SQL Server身份验证
    }
    
    # 创建导出器实例
    exporter = SQLServerSchemaExporter(**config)
    
    # 连接数据库
    if exporter.connect():
        try:
            # 导出到Excel
            output_file = exporter.export_to_excel()
            
            if output_file:
                print(f"\n导出完成!")
                print(f"文件位置: {os.path.abspath(output_file)}")
            
        except Exception as e:
            print(f"导出过程中出现错误: {str(e)}")
        finally:
            exporter.close()
    else:
        print("数据库连接失败，请检查连接参数")

if __name__ == "__main__":
    main()